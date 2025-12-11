import os
os.environ['OPENCV_VIDEOIO_DEBUG'] = '0'
os.environ['OPENCV_LOG_LEVEL'] = 'ERROR'

import cv2
import mediapipe as mp
import numpy as np
import csv
import json
import time
from flask import Flask, render_template, Response, jsonify, request, send_file
from flask_cors import CORS
from inference_sdk import InferenceHTTPClient
import threading
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import base64

# Vercel environment setup
VERCEL_ENV = os.getenv('VERCEL_ENV', 'development')

app = Flask(__name__, template_folder='templates', static_folder='static')
CORS(app)

# ========== CONFIGURATION ==========
WEIGHTS_FILE = '/tmp/model_art_weights.csv' if VERCEL_ENV == 'production' else 'model_art_weights.csv'
NAMES_FILE = '/tmp/model_art_names.csv' if VERCEL_ENV == 'production' else 'model_art_names.csv'
GESTURE_DATA_DIR = '/tmp/gesture_data' if VERCEL_ENV == 'production' else 'gesture_data'

# Create gesture data directory if not exists
os.makedirs(GESTURE_DATA_DIR, exist_ok=True)

ROBOFLOW_API_KEY = "5vFdrINv9FtzcV3AztHzI" 
ROBOFLOW_MODEL_ID = "object-dect-qhvpj/3" 
ROBOFLOW_URL = "https://detect.roboflow.com"

RHO = 0.90
ALPHA = 0.001
BETA = 1.0

# ========== FUZZY ART CLASS ==========
class FuzzyART:
    def __init__(self, rho=RHO, alpha=ALPHA, beta=BETA):
        self.rho = rho
        self.alpha = alpha
        self.beta = beta
        self.weights = []

    def complement_coding(self, x):
        x = np.clip(x, 0, 1)
        return np.concatenate((x, 1 - x))

    def classify(self, x_raw):
        if len(self.weights) == 0: 
            return -1
        x = self.complement_coding(x_raw)

        T = []
        for W_j in self.weights:
            fuzzy_and = np.minimum(x, W_j)
            T_j = np.sum(fuzzy_and) / (self.alpha + np.sum(W_j))
            T.append(T_j)
        
        j_star = np.argmax(T)
        W_jstar = self.weights[j_star]
        match_ratio = np.sum(np.minimum(x, W_jstar)) / np.sum(x)

        if match_ratio >= self.rho: 
            return j_star
        else: 
            return -2

# ========== MODEL FUNCTIONS ==========
def save_model_csv(art_model, gesture_names):
    with open(WEIGHTS_FILE, 'w', newline='') as f:
        writer = csv.writer(f)
        for w in art_model.weights:
            writer.writerow(w) 
    
    with open(NAMES_FILE, 'w', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(["id", "name"])
        for idx, name in gesture_names.items():
            writer.writerow([idx, name])

def load_model_csv():
    art = FuzzyART()
    names = {}
    
    if os.path.exists(WEIGHTS_FILE):
        with open(WEIGHTS_FILE, 'r') as f:
            reader = csv.reader(f)
            loaded_weights = []
            for row in reader:
                if not row: continue
                try:
                    w_vec = np.array(row, dtype=np.float32)
                    loaded_weights.append(w_vec)
                except ValueError:
                    pass
            art.weights = loaded_weights

    if os.path.exists(NAMES_FILE):
        with open(NAMES_FILE, 'r') as f:
            reader = csv.reader(f)
            next(reader, None)
            for row in reader:
                if len(row) >= 2:
                    try:
                        idx = int(row[0])
                        label = row[1]
                        names[idx] = label
                    except ValueError:
                        continue
    
    return art, names

def save_to_excel(gesture_name, features, confidence, timestamp=None):
    """Simpan gesture data real-time ke Excel"""
    try:
        excel_file = '/tmp/gestures_data.xlsx' if VERCEL_ENV == 'production' else 'gestures_data.xlsx'
        
        # Cek apakah file sudah ada
        if os.path.exists(excel_file):
            wb = load_workbook(excel_file)
            ws = wb.active
        else:
            # Buat workbook baru dengan header
            wb = Workbook()
            ws = wb.active
            ws.title = "Gestures"
            
            # Header styling
            headers = ["No", "Gesture Name", "Timestamp", "Confidence", "Features (Sample)"]
            ws.append(headers)
            
            # Style header
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Tambah data baru
        row_num = ws.max_row + 1
        timestamp = timestamp or datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        # Format features (ambil 5 pertama sebagai sample)
        if isinstance(features, (list, np.ndarray)):
            features_sample = ', '.join([f"{f:.3f}" for f in features[:5]]) + "..."
        else:
            features_sample = str(features)
        
        ws.append([row_num - 1, gesture_name, timestamp, f"{confidence:.2f}", features_sample])
        
        # Set column widths
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 35
        
        # Center alignment untuk beberapa kolom
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=4):
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Save file
        wb.save(excel_file)
        return True
    except Exception as e:
        print(f"[ERROR] Failed to save to Excel: {e}")
        return False

def extract_features(results):
    if not results.multi_hand_landmarks: 
        return None
    hand_landmarks = results.multi_hand_landmarks[0]
    base_x = hand_landmarks.landmark[0].x
    base_y = hand_landmarks.landmark[0].y
    feature_vector = []
    for lm in hand_landmarks.landmark:
        norm_x = (lm.x - base_x) + 0.5
        norm_y = (lm.y - base_y) + 0.5
        feature_vector.append(max(0.0, min(1.0, norm_x)))
        feature_vector.append(max(0.0, min(1.0, norm_y)))
    return np.array(feature_vector, dtype=np.float32)

# ========== GLOBAL STATE ==========
art, gesture_names = load_model_csv()
yolo_client = InferenceHTTPClient(api_url=ROBOFLOW_URL, api_key=ROBOFLOW_API_KEY)

mp_hands = mp.solutions.hands
mp_draw = mp.solutions.drawing_utils
hands = mp_hands.Hands(static_image_mode=False, max_num_hands=1, min_detection_confidence=0.5)

current_gesture = "Menunggu gambar dari browser..."
current_confidence = 0.0
current_features = None
auto_save_unknown = False
unknown_gesture_counter = 0
last_saved_gesture = None
lock = threading.Lock()

# ========== PROCESS FRAME FROM BROWSER ==========
def process_frame_data(frame_data):
    """Process frame yang dikirim dari browser"""
    global current_gesture, current_confidence, current_features
    global auto_save_unknown, unknown_gesture_counter, last_saved_gesture
    
    try:
        # Decode base64 image
        img_data = base64.b64decode(frame_data.split(',')[1])
        nparr = np.frombuffer(img_data, np.uint8)
        frame = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
        
        if frame is None:
            return None
        
        display_frame = frame.copy()
        
        status = "Mencari Tangan..."
        color = (255, 255, 0)
        vec = None
        cropped_frame = None
        confidence = 0.0

        # ROBOFLOW DETECTION
        try:
            predictions = yolo_client.infer(display_frame, model_id=ROBOFLOW_MODEL_ID, confidence=0.5)
            
            if predictions and 'predictions' in predictions:
                hand_predictions = [p for p in predictions['predictions'] if p['class'].lower() == 'hand']
                if hand_predictions:
                    best_pred = max(hand_predictions, key=lambda x: x['confidence'])
                    x_min, y_min, x_max, y_max = int(best_pred['x_min']), int(best_pred['y_min']), int(best_pred['x_max']), int(best_pred['y_max'])
                    
                    pad = 30 
                    H, W, _ = frame.shape
                    x_min = max(0, x_min - pad)
                    y_min = max(0, y_min - pad)
                    x_max = min(W, x_max + pad)
                    y_max = min(H, y_max + pad)
                    
                    cv2.rectangle(display_frame, (x_min, y_min), (x_max, y_max), (255, 0, 0), 2)
                    cropped_frame = frame[y_min:y_max, x_min:x_max]
                    confidence = best_pred.get('confidence', 0.0)
        except Exception as e:
            print(f"[ERROR] YOLO detection: {e}")
        
        # MEDIAPIPE PROCESSING
        frame_to_process = cropped_frame if cropped_frame is not None and cropped_frame.size > 0 else frame
        rgb_frame = cv2.cvtColor(frame_to_process, cv2.COLOR_BGR2RGB)
        results = hands.process(rgb_frame)

        if results.multi_hand_landmarks:
            mp_draw.draw_landmarks(display_frame, results.multi_hand_landmarks[0], mp_hands.HAND_CONNECTIONS)
            vec = extract_features(results)
            
            if vec is not None:
                idx = art.classify(vec)     
                
                if idx >= 0:
                    name = gesture_names.get(idx, f"Unknown ({idx})")
                    status = f"Gesture: {name}"
                    color = (0, 255, 0)
                    confidence = 0.95
                elif idx == -2:
                    status = "Tidak Dikenal"
                    color = (0, 0, 255)
                    confidence = 0.5
                    
                    if auto_save_unknown and last_saved_gesture != vec.tobytes():
                        try:
                            with lock:
                                gesture_name = f"unknown_{unknown_gesture_counter}"
                                filename, gesture_id = save_gesture_sample(gesture_name, vec)
                                unknown_gesture_counter += 1
                            last_saved_gesture = vec.tobytes()
                            status = f"Auto-saved: {gesture_name}"
                            color = (0, 165, 255)
                        except Exception as e:
                            print(f"[ERROR] Auto-save failed: {e}")

        with lock:
            current_gesture = status
            current_confidence = confidence
            current_features = vec.tolist() if vec is not None else None

        # UI
        cv2.rectangle(display_frame, (0, 0), (640, 90), (0, 0, 0), -1)
        cv2.putText(display_frame, status, (20, 40), cv2.FONT_HERSHEY_SIMPLEX, 1, color, 2)
        cv2.putText(display_frame, f"Confidence: {confidence:.2f} | Categories: {len(art.weights)}", 
                    (20, 75), cv2.FONT_ITALIC, 0.5, (200, 200, 200), 1)
        
        # Encode ke base64
        _, buffer = cv2.imencode('.jpg', display_frame, [cv2.IMWRITE_JPEG_QUALITY, 85])
        img_base64 = base64.b64encode(buffer).decode('utf-8')
        
        return img_base64
        
    except Exception as e:
        print(f"[ERROR] Frame processing failed: {e}")
        return None

# ========== ROUTES ==========
@app.route('/')
def index():
    return render_template('index.html')

@app.route('/api/process_frame', methods=['POST'])
def process_frame_endpoint():
    """Process frame dari browser webcam"""
    try:
        data = request.json
        frame_data = data.get('frame')
        
        if not frame_data:
            return jsonify({'status': 'error', 'message': 'No frame data'}), 400
        
        processed_image = process_frame_data(frame_data)
        
        if processed_image:
            with lock:
                return jsonify({
                    'status': 'success',
                    'image': processed_image,
                    'gesture': current_gesture,
                    'confidence': current_confidence,
                    'features': current_features,
                    'categories': len(art.weights)
                })
        else:
            return jsonify({'status': 'error', 'message': 'Frame processing failed'}), 500
            
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

def get_saved_gestures_count():
    """Hitung jumlah gesture yang tersimpan di folder gesture_data"""
    try:
        if os.path.exists(GESTURE_DATA_DIR):
            gesture_files = [f for f in os.listdir(GESTURE_DATA_DIR) if f.endswith('.json')]
            return len(gesture_files)
    except:
        pass
    return 0

@app.route('/api/gesture')
def get_gesture():
    with lock:
        saved_count = get_saved_gestures_count()
        return jsonify({
            'gesture': current_gesture,
            'confidence': current_confidence,
            'saved_gestures': saved_count,
            'features': current_features
        })

@app.route('/api/save_model', methods=['POST'])
def save_model():
    save_model_csv(art, gesture_names)
    return jsonify({'status': 'success', 'message': 'Model saved successfully'})

def save_gesture_sample(gesture_name, feature_vector):
    """Simpan gesture sample ke file JSON dan ART model"""
    global art, gesture_names
    
    timestamp = int(time.time() * 1000)
    filename = os.path.join(GESTURE_DATA_DIR, f"{gesture_name}_{timestamp}.json")
    
    if isinstance(feature_vector, np.ndarray):
        features_list = feature_vector.tolist()
        vec = feature_vector
    elif isinstance(feature_vector, list):
        features_list = feature_vector
        vec = np.array(feature_vector, dtype=np.float32)
    else:
        raise ValueError(f"Invalid feature_vector type: {type(feature_vector)}")
    
    data = {
        'gesture': gesture_name,
        'timestamp': timestamp,
        'features': features_list
    }
    
    with open(filename, 'w') as f:
        json.dump(data, f)
    
    # Train ke model ART
    x = np.clip(vec, 0, 1)
    x = np.concatenate((x, 1 - x))
    
    if len(art.weights) == 0:
        art.weights.append(x)
        idx = 0
    else:
        T = []
        for W_j in art.weights:
            fuzzy_and = np.minimum(x, W_j)
            T_j = np.sum(fuzzy_and) / (ALPHA + np.sum(W_j))
            T.append(T_j)
        
        sorted_indices = np.argsort(T)[::-1]
        idx = None
        
        for j_star in sorted_indices:
            W_jstar = art.weights[int(j_star)]
            match_ratio = np.sum(np.minimum(x, W_jstar)) / np.sum(x)
            
            if match_ratio >= RHO:
                new_weight = BETA * np.minimum(x, W_jstar) + (1 - BETA) * W_jstar
                art.weights[int(j_star)] = new_weight
                idx = int(j_star)
                break
        
        if idx is None:
            art.weights.append(x)
            idx = int(len(art.weights) - 1)
    
    gesture_names[int(idx)] = gesture_name
    save_model_csv(art, gesture_names)
    
    timestamp_str = datetime.fromtimestamp(timestamp / 1000).strftime("%Y-%m-%d %H:%M:%S")
    save_to_excel(gesture_name, features_list, 0.95, timestamp_str)
    
    return filename, int(idx)

@app.route('/api/save_gesture', methods=['POST'])
def save_gesture_endpoint():
    """Simpan gesture saat ini dengan nama yang diberikan"""
    global art, gesture_names
    
    data = request.json
    gesture_name = data.get('gesture_name', '').strip()
    feature_vector = data.get('features', None)
    
    if not gesture_name:
        return jsonify({'status': 'error', 'message': 'Nama gesture tidak boleh kosong'}), 400
    
    if feature_vector is None:
        return jsonify({'status': 'error', 'message': 'Feature vector tidak ditemukan'}), 400
    
    try:
        with lock:
            filename, gesture_id = save_gesture_sample(gesture_name, feature_vector)
        
        return jsonify({
            'status': 'success',
            'message': f'Gesture "{gesture_name}" berhasil disimpan',
            'filename': os.path.basename(filename),
            'gesture_id': gesture_id,
            'total_categories': len(art.weights)
        })
    
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/gestures/list', methods=['GET'])
def list_gestures():
    """List semua gesture yang sudah disimpan"""
    with lock:
        gestures = [
            {'id': int(idx), 'name': name}
            for idx, name in sorted(gesture_names.items())
        ]
    return jsonify({
        'status': 'success',
        'gestures': gestures,
        'total': len(gestures)
    })

@app.route('/api/auto_save/toggle', methods=['POST'])
def toggle_auto_save():
    """Toggle fitur auto-save untuk gesture yang tidak dikenal"""
    global auto_save_unknown
    auto_save_unknown = not auto_save_unknown
    status_msg = "ENABLED" if auto_save_unknown else "DISABLED"
    return jsonify({
        'status': 'success',
        'auto_save_enabled': auto_save_unknown,
        'message': f'Auto-save: {status_msg}'
    })

@app.route('/api/auto_save/status', methods=['GET'])
def get_auto_save_status():
    """Dapatkan status auto-save"""
    return jsonify({
        'status': 'success',
        'auto_save_enabled': auto_save_unknown,
        'unknown_counter': unknown_gesture_counter,
        'total_categories': len(art.weights)
    })

@app.route('/api/auto_save/reset', methods=['POST'])
def reset_auto_save_counter():
    """Reset counter unknown gesture"""
    global unknown_gesture_counter
    unknown_gesture_counter = 0
    return jsonify({
        'status': 'success',
        'message': 'Unknown gesture counter reset to 0'
    })

@app.route('/api/download/excel', methods=['GET'])
def download_excel():
    """Download Excel file dengan gesture data"""
    excel_file = '/tmp/gestures_data.xlsx' if VERCEL_ENV == 'production' else 'gestures_data.xlsx'
    if os.path.exists(excel_file):
        return send_file(excel_file, as_attachment=True, download_name='gestures_data.xlsx')
    else:
        return jsonify({'status': 'error', 'message': 'Excel file not found'}), 404

@app.route('/api/health', methods=['GET'])
def health_check():
    """Health check endpoint"""
    return jsonify({
        'status': 'healthy',
        'environment': VERCEL_ENV,
        'models': len(art.weights)
    })

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 8080))
    app.run(host='0.0.0.0', port=port, debug=True)

